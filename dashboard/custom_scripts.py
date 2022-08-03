from django.http import HttpResponse
from dns import resolver
import smtplib
import csv
from openpyxl import load_workbook,Workbook
import os
from functools import wraps
from django.http import HttpResponse
from users.models import *
from func_timeout import *
from bs4 import BeautifulSoup
import requests
import openai
from django.conf import settings as st
openai.api_key = st.OPENAI_API_KEY
import re
import tldextract

def requires_credit(view):
    @wraps(view)
    def _view(request, *args, **kwargs):
        sel_user_credit = user_credit.objects.get(user=request.user)
        if sel_user_credit.credits_remaining == 0:

            return HttpResponse("You don't have enough credits to perform this action")
        return view(request, *args, **kwargs)
    return _view


#function for getting mx mx records of a single domain (no input email)
@func_set_timeout(5)
def get_mx_records_domain(domain):
    mx_records = []
    try:
        answers = resolver.resolve(domain, 'MX')
        for rdata in answers:
            mx_records.append(rdata.to_text().split(' ')[1].rstrip('.'))
    except:
        pass
    
    return mx_records


# function for getting mx records of a email domain
@func_set_timeout(5)
def get_mx_records(email):
    domain = email.split('@')[1]
    mx_records = []
    try:
        answers = resolver.resolve(domain, 'MX')
        for rdata in answers:
            mx_records.append(rdata.to_text().split(' ')[1].rstrip('.'))
    except:
        pass
    
    return mx_records

#check if the email is valid
@func_set_timeout(5)
def is_valid_email(mx_record,email):

    try:

        server = smtplib.SMTP()
        # uncomment the below line if you want to see full output.
        #server.set_debuglevel(1)

        #This is just a fake email that doesn't probably exist for smtp.mail(fromAddress)  
        fromAddress = 'shakibulislam@gmail.com'

        # SMTP Conversation
        server.connect(mx_record)
        server.helo(server.local_hostname) ### server.local_hostname(Get local server hostname)
        server.mail(fromAddress)
        code, message = server.rcpt(str(email))
        server.quit()

        

        # Assume SMTP response 250 is success
        if (code == 250 or code == 251 or code == 220 or code == 200):
            return True
        else:
            return False
    
    except:
        return False



#xlsx file reading and writing starts here


def xlsx_info(filepath):

	information = {}
	information['column_names'] =[]

	wb = load_workbook(filepath)

	wc = wb.active

	check_rows = wc.rows

	check_cols = wc.columns

	#storing the count of first row total
	for row in tuple(check_rows):
		information['total_columns']=len(row)
		break

	#total rows count for this file including first row
	information['total_rows']=(len(tuple(wc.rows)))

	#getting the first row names / column names as first row is always column header

	get_first_row = tuple(wc.rows)[0]

	for item in list(get_first_row):
		
		information['column_names'].append(item.value)
	


	return information




def xlsx_write_on_new_column(curr_row_num,column_num_total,value,filepath,is_new_file = False):
    try:
        wb=None

        if is_new_file == False:
            wb = load_workbook(filepath)

        else:
            wb = Workbook()
            
        wc = wb.active
        wc.cell(column=column_num_total+1,row=curr_row_num,value=value)
        wb.save(filepath)
        return True
        
    except:
        return False




#create function for creating and writing on xlsx file using openpyxl
def xlsx_create_and_write(filepath,datalist,headerlist):
    try:
        wb=Workbook()
        wc = wb.active
        for header in range(1,len(headerlist)+1):
            wc.cell(column=header,row=1,value=headerlist[header-1])

        for row in range(2,len(datalist)+2):
            for column_id in range(1,len(headerlist)+1):
                wc.cell(column=column_id,row=row,value=datalist[row-2][column_id-1])

        
        try:
            wb.save(filepath)

        except:
            

            wb.save(filepath)

        return True

    except:
        return False
            



#create method retrive columns

def xlsx_retrive_column_data(row,column,filepath):
	try:
		wb = load_workbook(filepath)

		wc = wb.active

		val = wc.cell(row=row,column=column).value

		return val

	except:
		return False





#csv to xlsx conversion function
def csv_to_xlsx(filepath,file_name):

   # new_path = filepath[:len(filepath)-3]+"xlsx"
    file_name = file_name[:len(file_name)-3].split('/')[-1]
    new_path = file_name+'xlsx'

    
    wb = Workbook()
    wc = wb.active
    rows = []
    try:
        with open(filepath) as csv_file:
            csv_inst = csv.reader(csv_file)
            for row in csv_inst:
                rows.append(row)
                
        for ro in range(1,len(rows)+1):
            sel_row = rows[int(ro)-1]
            counter = 1
            for item in sel_row:
                wc.cell(row=ro,column=counter,value=item)
                counter+=1
                
        wb.save('media/uploads/'+new_path)
        return True
    except:
        return False



def get_author_name(url):
    get_contents = requests.get(url)
    soup = BeautifulSoup(get_contents.content,'lxml')
    all_a_tags = soup.find_all('a',href=True)
    
    for link in all_a_tags:
        lnk = link["href"]
        if any(["author" in lnk,"Author" in lnk,"contributor" in lnk,"profile" in lnk,"user" in lnk,"writer" in lnk]):
            print("Author name is: "+link.string.strip())
            return {'string':link.string.strip(),'source_code':get_contents.content}
            break

    print("Author name not found")

    return {'string':None,'source_code':get_contents.content}
        
    






def extract_names(text):
    astring = text

    

    response_name = openai.Completion.create(
	  model="text-davinci-002",
	  prompt=f"find all the human names from the following text and separate names by comma: {astring}",
	  temperature=0.7,
	  max_tokens=35,
	  top_p=1,
	  frequency_penalty=0,
	  presence_penalty=0
	)


    return str(response_name.choices[0].text).strip()


def get_author_name_extracting_attribute(url,source_code=None):

    if source_code == None:
        get_contents = requests.get(url)

        soup = BeautifulSoup(get_contents.content,'lxml')
    else:
        soup = BeautifulSoup(source_code,'lxml')
        
    find_all_class_containg_author = soup.find_all(class_=re.compile(r"author|writer"))

	#string=re.compile(r"Written|article by|author|writer|Writer|written|editor|Author|published|ARTICLE|article|Article|By|by"))
    # 
    
    list_of_possible_authors = []
    for el in find_all_class_containg_author:
        if el.string !=  None and len(el.string) < 30:
            if el.string not in list_of_possible_authors and len(el.string.split(" ")) < 10:
                list_of_possible_authors.append(el.string)
                print('hello author ',el.string,'\n')

	

    list_of_possible_ele = []

    for ele in soup.h1.find_next_siblings()[:4]:
        for k in ele.descendants:
            
            if k.string != None and len(k.string) < 20 and len(k.string) > 7 and k.string != "\n":
                list_of_possible_ele.append(k.string)
                print(k.string,'\n')

    print('ele is ',list_of_possible_ele)

    

    list_of_possible_authors = set(list_of_possible_authors + list_of_possible_ele)
    
    unique_author_list_without_validation = list(list_of_possible_authors)

    unique_author_list = []

    for au in unique_author_list_without_validation:
        if len(au.split(" ")) > 1:
            unique_author_list.append(au)


    
    st = str(unique_author_list)[1:-1].strip().replace("'",'')

    
    
    names_in_comma = extract_names(st)

    #print('all_names are ', list_of_possible_ele)


    print('possible human names are -- ',names_in_comma)
    
    if len(names_in_comma.split(',')) != 0:
        return names_in_comma.split(',')
        
    else:
        return None

	


